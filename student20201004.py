import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


wb1 = load_workbook(filename="student.xlsx")
ws1 = wb1.active

Rmax = ws1.max_row
Cmax = ws1.max_column
stu = Rmax - 2

for row in range(2, Rmax + 1):
    ws1.cell(row=row, column=7, value=ws1.cell(row,3).value*0.3 + ws1.cell(row, 4).value*0.35 + ws1.cell(row, 5).value*0.34 + ws1.cell(row, 6).value*1)

wb1.save("student.xlsx")

data=pd.read_excel('student.xlsx') 
data=data.sort_values(by='total', ascending=False)

with pd.ExcelWriter('student.xlsx') as writer:
  data.to_excel(writer,sheet_name="sheet1",index=False) 

wb2 = load_workbook(filename="student.xlsx")
ws2 = wb2.active

cnt=0
for row in range(2, Rmax + 1):
    if ws2.cell(row, 7).value < 40:
        ws2.cell(row, 8, value="F")
        continue

    if cnt <= stu * 0.3:
        if cnt <= stu * 0.3 * 0.5:
            ws2.cell(row, 8, value="A+")
        else:
            ws2.cell(row, 8, value="A0")
    elif cnt <= stu * 0.7:
        if cnt <= stu * 0.3 + stu * 0.4 * 0.5:
            ws2.cell(row, 8, value="B+")
        else:
            ws2.cell(row, 8, value="B0")
    else:
        if cnt <= stu * 0.7 + stu * 0.3 * 0.5:
            ws2.cell(row, 8, value="C+")
        else:
            ws2.cell(row, 8, value="C0")

    cnt+=1


wb2.save("student.xlsx")

data=pd.read_excel('student.xlsx') 
data=data.sort_values(by='id')

with pd.ExcelWriter('student.xlsx') as writer:
    data.to_excel(writer,sheet_name="sheet1",index=False) 